__author__ = 'denisthamrin'
data_folder = '../project2_data/[!~$]*.xlsx'
filename = '../project2_data/Ascot-Vale-Suburb - XLSX.xlsx'


import glob
from openpyxl import load_workbook
from sklearn import manifold
import numpy as np
import matplotlib.pyplot as plt
import json

avail_languages = {}

#Cell Data
convert_to_cell_func = lambda x:  "C" + str(x)
countries_cell = map(convert_to_cell_func,range(181, 196))
languages_cell = map(convert_to_cell_func, range(196, 211))
aborigin_percentage_cell = "C172"
overseas_percentage_cell = "C174"
born_in_non_english_percentage_cell = "C176"
lote_percentage_cell = "C178"
poor_english_cell = "C180"

unique_countries = {}
unique_languages = {}
suburb_list = []
def process_excels():
    files = glob.glob(data_folder)

    feature_list = []

    for excel in files:
        suburb = excel.split('-Suburb')[0].split('/')[2]
        suburb_list.append(suburb)
        curr_wb = load_workbook("%s" % (excel ))
        curr_ws = curr_wb["data"]


# Top 5 language and contries
        countries = process_percentage(countries_cell, curr_ws)
        languages = process_percentage(languages_cell, curr_ws)

        # print suburb + " " + countries[0]
        # for c in countries:
        #     if c not in unique_countries:
        #         unique_countries[c] = 1
        #     else:
        #         unique_countries[c] += 1
        #
        #
        # for l in languages:
        #     if l not in unique_languages:
        #         unique_languages[l] = 1
        #     else:
        #         unique_languages[l] += 1



    # print json.dumps(unique_countries)
    # print "\\"
    # print json.dumps(unique_languages)


def process_percentage(data_rows,ws):
    # [(language/countries,percentage)]
    unique = []
    i = 0

    for row in data_rows:
        if i == 2:

            unique.append((language.encode('ascii')))
            i = 0
            continue
        elif i == 0:
            language = ws[row].value


        i += 1
    return unique

process_excels()

plt.bar(range(len(unique_languages)), unique_languages.values(), align='center')
plt.xticks(range(len(unique_languages)), unique_languages.keys(), rotation=25)
plt.show()