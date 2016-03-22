__author__ = 'denisthamrin'
data_folder = '../project2_data/[!~$]*.xlsx'
filename = '../project2_data/Ascot-Vale-Suburb - XLSX.xlsx'

import glob
from openpyxl import load_workbook

from sklearn.cluster import KMeans
from sklearn import manifold
import numpy as np
import matplotlib.pyplot as plt
import mpld3
from mpld3 import plugins
import pickle

avail_languages = {}

def convert_countries_to_numeric(x):
    x = x.lower()
    if x not in avail_languages:
        avail_languages[x] = convert_countries_to_numeric.counter
        convert_countries_to_numeric.counter += 1
    return avail_languages[x]
convert_countries_to_numeric.counter = 1

def convert_percentage_to_numeric(x):
    if x == "n/a":
        return 0.0
    else:
        return float(x)

# given rows and worksheet extract value (work for diversity(language and countries for now)

def process_percentage(data_rows,ws):
    # [(language/countries,percentage)]
    percentage_list = []
    i = 0

    for row in data_rows:
        if i == 2:

            percentage = convert_percentage_to_numeric(ws[row].value)
            percentage_list.append((language, percentage))
            i = 0
            continue
        elif i == 0:
            language = ws[row].value
            language = convert_countries_to_numeric(language)


        i += 1
    return percentage_list




#Cell Data
convert_to_cell_func = lambda x:  "C" + str(x)
countries_cell = map(convert_to_cell_func,range(181, 196))
languages_cell = map(convert_to_cell_func, range(196, 211))
aborigin_percentage_cell = "C172"
overseas_percentage_cell = "C174"
born_in_non_english_percentage_cell = "C176"
lote_percentage_cell = "C178"
poor_english_cell = "C180"


suburb_list = []

def custom_distance(X,Y):
    distance = []

    for v1 in X:
        curr_distance = []

        for v2 in Y:
            curr_distance.append(measure(v1, v2))

        distance.append(curr_distance)


    return distance


def measure(v1, v2):
    custom_vector = []
    for i in range(0,6):
        custom_vector.append(countries_distance(v1[i], v2[i]) * max((5-i*2),1/2) )

    custom_vector = reduce(lambda a, b: (a+b), custom_vector)
    return custom_vector


def language_distance(x,y):
    if x == y:
        return 0.0
    else:
        return 1.0


def countries_distance(x,y):
    if x == y:
        return 0.0
    else:
        return 1.0

def process_excels():
    files = glob.glob(data_folder)

    feature_list = []

    for excel in files:
        suburb = excel.split('-Suburb')[0].split('/')[2]
        suburb_list.append(suburb)
        curr_wb = load_workbook("%s" % (excel ))
        curr_ws = curr_wb["data"]


        # Top 5 language and contries
        countries_origin_percentage = process_percentage(countries_cell, curr_ws)
        language_percentage = process_percentage(languages_cell, curr_ws)
        get_first_element_func = lambda x: x[0]
        countries = map(get_first_element_func, countries_origin_percentage)
        languages = map(get_first_element_func, language_percentage)

        #other feature
        abo_percentage = curr_ws[aborigin_percentage_cell].value
        ove_percentage = curr_ws[overseas_percentage_cell].value
        born_in_non_english_percentage = curr_ws[born_in_non_english_percentage_cell].value
        lote_percentage = curr_ws[lote_percentage_cell].value
        poor_english = curr_ws[poor_english_cell].value

        aggregated_feature = map(convert_percentage_to_numeric,
                          [abo_percentage,ove_percentage,born_in_non_english_percentage,lote_percentage,poor_english])



        feature = countries + languages + aggregated_feature

        feature_list.append(tuple(feature))
    return feature_list


def plot_d3js():
    # Define some CSS to control our custom labels
    css = """
    table
    {
      border-collapse: collapse;
    }
    th
    {
      color: #ffffff;
      background-color: #000000;
    }
    td
    {
      background-color: #cccccc;
    }
    table, th, td
    {
      font-family:Arial, Helvetica, sans-serif;
      border: 1px solid black;
      text-align: right;
    }
    """

    fig, ax = plt.subplots()
    ax.grid(True, alpha=0.3)
    labels = suburb_list
    x = coords[:, 0]
    y = coords[:, 1]

    points = ax.plot(x, y, 'o', color='b',
                     mec='k', ms=15, mew=1, alpha=.6)

    ax.set_xlabel('x')
    ax.set_ylabel('y')
    ax.set_title('Ethnicity', size=20)

    tooltip = plugins.PointHTMLTooltip(points[0], labels,
                                       voffset=10, hoffset=10, css=css)
    plugins.connect(fig, tooltip)

    mpld3.show()

def plot_matplotlib():
    plt.subplots_adjust(bottom = 0.1)
    plt.scatter(
        coords[:, 0], coords[:, 1], marker = 'o'
        )
    for label, x, y in zip(suburb_list, coords[:, 0], coords[:, 1]):
        plt.annotate(
            label,
            xy = (x, y), xytext = (-20, 20),
            textcoords = 'offset points', ha = 'right', va = 'bottom',
            bbox = dict(boxstyle = 'round,pad=0.5', fc = 'yellow', alpha = 0.5),
            arrowprops = dict(arrowstyle = '->', connectionstyle = 'arc3,rad=0'))

    plt.show()

def plot(label,xy,xytext):
    plt.annotate(
        label,
        xy = xy, xytext = xytext,
        textcoords = 'offset points', ha = 'right', va = 'bottom',
        bbox = dict(boxstyle = 'round,pad=0.5', fc = 'white', alpha = 0.5),
        arrowprops = dict(arrowstyle = '->', connectionstyle = 'arc3,rad=0'))


def plot_manual_label():

    suburb_coord_dict = {}
    for label, x, y in zip(suburb_list, coords[:, 0], coords[:, 1]):
        suburb_coord_dict[label] = (x,y)

    axis = [(0,-20),(80,20),(-20,40),(80,-20),(20,-20),(0,-60),(80,-60)]
    upleft = ["Fawkner","Waterways","Port-Melbourne","Sorrento","Prahran","Springvale","Parkville","Noble-Park","Glenroy"]
    upright = ["Mordialloc","St-Andrews-Beach","Northcote","Malvern","Craigieburn","Croydon","St-Kilda-East","Moorabin","Melbourne-Airport","Footscray"]
    downright = ["Pascoe-Vale-South","North-Melbourne","South-Melbourne","Somerville","Braybrook","Malvern-East","Windsor","Murrumbeena","Tyabb"]
    downleft = ["Moorabbin"]
    down = ["Ascot-Vale"]
    longdownleft = ["St-Kilda-West"]
    longdownright = ["Toorak","St-Kilda"]
    print "printing"
    for s in suburb_coord_dict:
        if s in upleft:
            plot(s,suburb_coord_dict[s],axis[2])
        elif s in upright:
            plot(s,suburb_coord_dict[s],axis[1])
        elif s in downright:
            plot(s,suburb_coord_dict[s],axis[3])
        elif s in downleft:
            plot(s,suburb_coord_dict[s],axis[0])
        elif s in longdownleft:
            plot(s,suburb_coord_dict[s],axis[5])
        elif s in longdownright:
            plot(s,suburb_coord_dict[s],axis[6])

        elif s in down:
            plot(s,suburb_coord_dict[s],axis[4])


def plot_kmean():
    n_clusters = 8
    k_means = KMeans(init='k-means++', n_clusters=n_clusters, n_init=10)
    X = coords
    k_means.fit(X)
    k_means_labels = k_means.labels_
    k_means_cluster_centers = k_means.cluster_centers_

    colors = ['#4EACC5', '#FF9C34', '#4E9A06',"#F0F8FF","#800080","#A52A2A","#556B2F","#FA8072","#D2691E"]
    for k, col in zip(range(n_clusters), colors):
        my_members = k_means_labels == k
        cluster_center = k_means_cluster_centers[k]
        plt.plot(X[my_members, 0], X[my_members, 1], 'w',
                markerfacecolor=col, marker='o', markersize=30)
        plt.plot(cluster_center[0], cluster_center[1], 'x', markerfacecolor=col,
                markeredgecolor='red', markersize=40)


feature_list = process_excels()
feature_list = custom_distance(feature_list, feature_list)
mds = manifold.MDS(dissimilarity='precomputed')
coords = mds.fit(feature_list)
coords = coords.embedding_
plot_kmean()
plot_manual_label()
plt.show()

#
# # pickle.dump( coords, open( "coords.p", "wb" ) )
# # pickle.dump( suburb_list, open( "suburb_list.p", "wb" ) )
#
# coords = pickle.load( open( "coords.p", "rb" ) )
# suburb_list =  pickle.load( open( "suburb_list.p", "rb" ) )
