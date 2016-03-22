__author__ = 'viva'
import parser
import pickle
import os
from openpyxl import load_workbook


scopes = {
    'community': [0, 2],
    'geography': [2, 17],
    'land_use': [17, 27],
    'population_2012': [27, 52],
    'population_2007': [52, 77],
    'population_change': [77, 90],
    'services': [90, 114],
    'socio_demographic': [114, 170],
    'diversity': [170, 210],
    'hospital': [210, 226]
}

#data_path = os.getcwd() + os.sep + 'project2_data'
data_path = os.pardir + os.sep + 'project2_data'


# generate profile by category
def features_by_category(category):
    index = 0

    features_of_all_suburbs = {}
    for suburb_file in os.listdir(data_path):

        try:

            suburb = load_workbook(data_path + os.sep + suburb_file)

            scope = scopes[category]
            feature_names = suburb['data'].columns[1][scope[0]:scope[1]]
            feature_values = suburb['data'].columns[2][scope[0]:scope[1]]

            features = {}
            for feature_name, feature_value in zip(feature_names, feature_values):
                features[feature_name.value] = feature_value.value

            suburb_name = suburb_file.split('-Suburb')[0].replace('-', '_')
            features_of_all_suburbs[suburb_name] = features

        except:
            pass

    return features_of_all_suburbs


def get_features():
    features_of_all_suburbs = {}

    for suburb_file in os.listdir(data_path):

        try:

            suburb = load_workbook(data_path + os.sep + suburb_file)
            feature_names = suburb['data'].columns[1][0:226]
            feature_values = suburb['data'].columns[2][0:226]

            features = {}
            for feature_name, feature_value in zip(feature_names, feature_values):
                features[feature_name.value] = feature_value.value

            suburb_name = suburb_file.split('-Suburb')[0].replace('-', '_')
            features_of_all_suburbs[suburb_name] = features

        except:
            pass

    return features_of_all_suburbs


def dump(category):
    f = open(category, 'wb')
    pickle.dump(features_by_category(category),f)
    f.close()

def dump_all():
    f = open('feature_all', 'wb')
    pickle.dump(get_features(),f)
    f.close()

dump('socio_demographic')

dump_all()
